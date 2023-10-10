import os
import zipfile
from PyPDF2 import PdfReader
from openpyxl import load_workbook
import pandas

path_root = os.path.dirname(os.path.abspath(__file__))
path_tmp = os.path.join(path_root, 'tmp')
path_resources = os.path.join(path_root, 'resources')
path_resources_zip_file = os.path.join(path_root, 'resources', 'file.zip')
path_tmp_zip_file = os.path.join(path_root, 'tmp', 'file.zip')


def test_create_zip():
    '''
    Создаем архив
    '''
    os.chdir(path_resources)
    with zipfile.ZipFile('file.zip', 'w') as z:
        z.write('Hello.txt')
        z.write('file_example_XLS_10.xls')
        z.write('file_example_XLSX_50.xlsx')
        z.write('Python Testing with Pytest (Brian Okken).pdf')
    if not os.path.isdir(path_tmp):
        os.mkdir(path_tmp)
    if os.path.exists(path_tmp_zip_file):
        os.remove(path_tmp_zip_file)
    os.rename(path_resources_zip_file, path_tmp_zip_file)


def test_file_txt():
    '''
    Проверяем содержимое файла txt
    '''
    with zipfile.ZipFile(path_tmp_zip_file) as zip:
        with zip.open('Hello.txt') as txt:
            assert txt.read().decode('utf-8') == 'Hello world!'


def test_file_xls():
    '''
    Проверяем размер и один из заголовков файла xls
    '''
    with zipfile.ZipFile(path_tmp_zip_file, 'r') as zip:
        with zip.open("file_example_XLS_10.xls") as xls:
            name_header = pandas.read_excel(xls).head()
            assert 'First Name' in name_header
            xls_file_size = zip.infolist()[1].file_size
            assert xls_file_size == 8704


def test_file_xlsx():
    '''
    Проверяем содержимое ячейки и размер файла xlsx
    :return:
    '''
    with zipfile.ZipFile(path_tmp_zip_file) as zip:
        with zip.open("file_example_XLSX_50.xlsx") as xlsx:
            workbook = load_workbook(xlsx)
            sheet = workbook.active
            sheet = sheet.cell(row=2, column=2).value
            print(sheet)
            assert sheet == 'Dulce'
            xlsx_file_size = zip.infolist()[2].file_size
            assert xlsx_file_size == 7360


def test_file_pdf():
    '''
    Проверяем количество страниц и размер файла pdf
    '''
    with zipfile.ZipFile(path_tmp_zip_file) as zip:
        with zip.open('Python Testing with Pytest (Brian Okken).pdf') as pdf:
            pdf_data = PdfReader(pdf)
            number_of_page = len(pdf_data.pages)
            assert number_of_page == 256
            pdf_file_size = zip.infolist()[3].file_size
            assert pdf_file_size == 3035139
